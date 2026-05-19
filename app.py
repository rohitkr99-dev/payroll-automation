import streamlit as st
import pandas as pd

from datetime import datetime, time
from io import BytesIO

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

st.title("Payroll Automation System")

uploaded_file = st.file_uploader(
    "Upload attendance file",
    type=["xls", "xlsx"]
)

def process_payroll(file):

    raw = pd.read_excel(file, header=None)

    employees = []

    current_emp = None

    all_dates = set()

    # READ RAW DATA
    for _, row in raw.iterrows():

        first = row[0]

        # EMPLOYEE ROW
        if isinstance(first, str) and "Mr." in first:

            parts = first.strip().split(" ", 1)

            emp_id = parts[0]

            emp_name = parts[1] if len(parts) > 1 else ""

            current_emp = {
                "Employee ID": emp_id,
                "Employee Name": emp_name,
                "days": {}
            }

            employees.append(current_emp)

            continue

        # DATE ROW
        if current_emp and isinstance(first, str) and "/" in first:

            try:

                dt = datetime.strptime(
                    first.strip(),
                    "%d/%b/%Y"
                ).date()

            except:

                continue

            punches = []

            for col in [1, 2, 3, 4]:

                val = row[col]

                if pd.notna(val):

                    try:

                        punches.append(
                            pd.to_datetime(val).time()
                        )

                    except:
                        pass

            current_emp["days"][dt.day] = punches

            all_dates.add(dt.day)

    sorted_dates = sorted(all_dates)

    summary_rows = []

    issue_rows = []

    # PROCESS EMPLOYEES
    for emp in employees:

        for row_type in ["NH", "FOT", "EOT", "DETAILS"]:

            row = {
                "Employee ID": emp["Employee ID"],
                "Employee Name": emp["Employee Name"],
                "Type": row_type
            }

            for d in sorted_dates:

                punches = emp["days"].get(d, [])

                value = ""

                if len(punches) > 0:

                    first_in = punches[0]

                    last_out = punches[-1]

                    issues = []

                    # SINGLE PUNCH HANDLING
                    if len(punches) == 1:

                        if first_in < time(12, 0):

                            issues.append(
                                f"Single Morning Punch {first_in.strftime('%H:%M')}"
                            )

                            last_out = time(18, 0)

                        else:

                            issues.append(
                                f"Single Evening Punch {first_in.strftime('%H:%M')}"
                            )

                            first_in = time(8, 0)

                    # EARLY IN RULE
                    if first_in <= time(7, 0):

                        issues.append(
                            f"Early IN {first_in.strftime('%H:%M')}"
                        )

                        first_in = time(8, 0)

                    # EXTRA OT RULE
                    eot = 0

                    out_minutes = (
                        last_out.hour * 60 +
                        last_out.minute
                    )

                    if out_minutes >= 19 * 60:

                        eot = int(
                            (out_minutes - 18 * 60) // 60
                        )

                    # ROW TYPE VALUES
                    if row_type == "NH":

                        value = 8

                    elif row_type == "FOT":

                        value = 1

                    elif row_type == "EOT":

                        value = eot

                    elif row_type == "DETAILS":

                        value = (
                            f"{first_in.strftime('%H:%M')} - "
                            f"{last_out.strftime('%H:%M')}"
                        )

                        if issues:

                            value += " | " + "; ".join(issues)

                            issue_rows.append({
                                "Employee ID": emp["Employee ID"],
                                "Employee Name": emp["Employee Name"],
                                "Date": d,
                                "Issue": "; ".join(issues)
                            })

                row[str(d)] = value

            summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)

    issues_df = pd.DataFrame(issue_rows)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # WRITE SHEETS
        summary_df.to_excel(
            writer,
            sheet_name="Payroll Summary",
            index=False
        )

        issues_df.to_excel(
            writer,
            sheet_name="Collective Issues",
            index=False
        )

        workbook = writer.book

        # COLORS
        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        nh_fill = PatternFill(
            start_color="E2F0D9",
            end_color="E2F0D9",
            fill_type="solid"
        )

        fot_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        eot_fill = PatternFill(
            start_color="FCE4D6",
            end_color="FCE4D6",
            fill_type="solid"
        )

        detail_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )

        issue_fill = PatternFill(
            start_color="F4CCCC",
            end_color="F4CCCC",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # FORMAT SUMMARY SHEET
        ws = writer.sheets["Payroll Summary"]

        ws.freeze_panes = "D2"

        for row in ws.iter_rows():

            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # HEADER STYLE
        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

        # ROW COLORS
        for row_num in range(2, ws.max_row + 1):

            row_type = ws.cell(
                row=row_num,
                column=3
            ).value

            fill = None

            if row_type == "NH":
                fill = nh_fill

            elif row_type == "FOT":
                fill = fot_fill

            elif row_type == "EOT":
                fill = eot_fill

            elif row_type == "DETAILS":
                fill = detail_fill

            if fill:

                for col_num in range(1, ws.max_column + 1):

                    ws.cell(
                        row=row_num,
                        column=col_num
                    ).fill = fill

        # AUTO WIDTH
        for column_cells in ws.columns:

            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )

            adjusted_width = min(length + 4, 40)

            ws.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = adjusted_width

        # ROW HEIGHTS
        for row_num in range(1, ws.max_row + 1):

            ws.row_dimensions[row_num].height = 28

        # FORMAT ISSUES SHEET
        ws2 = writer.sheets["Collective Issues"]

        ws2.freeze_panes = "A2"

        for row in ws2.iter_rows():

            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # ISSUES HEADER
        for cell in ws2[1]:

            cell.fill = header_fill

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

        # ISSUE ROW COLORS
        for row_num in range(2, ws2.max_row + 1):

            for col_num in range(1, ws2.max_column + 1):

                ws2.cell(
                    row=row_num,
                    column=col_num
                ).fill = issue_fill

        # ISSUES WIDTH
        for column_cells in ws2.columns:

            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )

            adjusted_width = min(length + 5, 50)

            ws2.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = adjusted_width

        # ISSUES ROW HEIGHTS
        for row_num in range(1, ws2.max_row + 1):

            ws2.row_dimensions[row_num].height = 28

    output.seek(0)

    return output

# MAIN APPLICATION
if uploaded_file:

    st.success("File uploaded successfully")

    try:

        result = process_payroll(uploaded_file)

        st.download_button(
            label="Download Payroll Report",
            data=result,
            file_name="payroll_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        st.error(
            f"Error processing file: {str(e)}"
        )
